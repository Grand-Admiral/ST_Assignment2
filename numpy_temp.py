import numpy
import openpyxl as pyxl
import sqlite3 as sql3
import matplotlib.pyplot as plt

conn = sql3.connect("data/collectedData.db")
cur = conn.cursor()

Data = []
Country = []

#create exel reader
dest_filename = "worldTemperature.xlsx"
wb = pyxl.load_workbook(dest_filename)
Name = wb.get_sheet_names()
print("sheet 1 got")

name = [Name]
print(name)

ws = wb.active
if Name == "Comparison":
    print("already have")
else:
    ws = wb.create_sheet("Comparison")

sheet = wb.get_sheet_by_name('Comparison')
sheet.cell(row = 1,column = 1, value = "year")
sheet.cell(row = 1,column = 2, value = "Temp")
sheet.cell(row = 1,column = 3, value = "State")

sheet.cell(row = 1,column = 5, value = "year")
sheet.cell(row = 1,column = 6, value = "Temp")
sheet.cell(row = 1,column = 7, value = "Country")

TMP = []

def SQLStateAVG():
    global Data,TMP
    taks1 = cur.execute("Select distinct State from State where Country == 'Australia';")# select all cities in china
    v = 2
    for row in taks1:
        TMP.append(row[0])
        
    for year in range(1852,2014):#year
        for i in TMP: #states ACT
            taks1 = cur.execute("Select distinct avg(AverageTemp) as Temp, State from State where Date Like '"+str(year)+"%' and Country == 'Australia' and State == '"+str(i)+"';")
            for row in taks1:#data
               if row[0] != None:
                   print("Year: ", year)
                   print("Avg Temp= ", row[0],)
                   print("State = ", row[1],"\n")
                   
                   sheet.cell(row = v,column = 1, value = year)
                   sheet.cell(row = v,column = 2, value = row[0])
                   sheet.cell(row = v,column = 3, value = row[1])
                   v+=1


               for i in [1,0]:
                   Data.append(row[i])
            


def SQLCountryAVG():
    global Country
    v = 2
    for l in range(1852,2014):#Date
        taks1 = cur.execute("Select avg(AverageTemp)as Temp, Country from Country where Date Like '"+str(l)+"%' and Country == 'Australia';")# select all cities in china
        for row in taks1:#data
           #print("Date = ", row[0],)
           print("Year: ", l)
           print("Avg Temp= ", row[0],)
           print("Country = ", row[1],"\n")
           Country.append(row[0])

           sheet.cell(row = v,column = 5, value = l)
           sheet.cell(row = v,column = 6, value = row[0])
           sheet.cell(row = v,column = 7, value = row[1])
           v+=1


def Diff():
    print("################################")
    global Data, Country,TMP
    count = [int(0),int(1)]#counter to track location in the list
    State1 = []
    State2 = []
    State3 = []
    State4 = []
    State5 = []
    State6 = []
    State7 = []
    State8 = []
    
    for s in Country:#only used once can help keep track of the many
        print(count)
        for i in TMP:#states
            print(i)
            try:#try to find all data and sift it into its respective category
                print("Difference between state and country data",float(Data[count[1]]) - float(Country[count[0]]))
                if i == "Queensland":
                    State1.append(float(Data[count[1]]) - float(Country[count[0]]))
                elif i == "Australian Capital Territory":
                    State2.append(float(Data[count[1]]) - float(Country[count[0]]))
                elif i == "New South Wales":
                    State3.append(float(Data[count[1]]) - float(Country[count[0]]))
                elif i == "Northern Territory":
                    State4.append(float(Data[count[1]]) - float(Country[count[0]]))
                elif i == "South Australia":
                    State5.append(float(Data[count[1]]) - float(Country[count[0]]))
                elif i == "Tasmania":
                    State6.append(float(Data[count[1]]) - float(Country[count[0]]))
                elif i == "Victoria":
                    State7.append(float(Data[count[1]]) - float(Country[count[0]]))
                elif i == "Western Australia":
                    State8.append(float(Data[count[1]]) - float(Country[count[0]]))
            except:
                print("no data")
                if i == "Queensland":
                    State1.append(0)
                elif i == "Australian Capital Territory":
                    State2.append(0)
                elif i == "New South Wales":
                    State3.append(0)
                elif i == "Northern Territory":
                    State4.append(0)
                elif i == "South Australia":
                    State5.append(0)
                elif i == "Tasmania":
                    State6.append(0)
                elif i == "Victoria":
                    State7.append(0)
                elif i == "Western Australia":
                    State8.append(0)
            count[1] += 2
        count[0] += 1
    
    years = [] #create the years list for the chart
    for i in range(1852, 2014):
        years.append(i)
    # generate difference chart
    plt.plot(years,State1,years,State2,years,State3,years,State4,years,State5,years,State6,years,State7,years,State8)
    plt.xlabel('Years')
    plt.ylabel('Temp difference C')
    plt.title("Difference between state and Country")
    plt.show() 


###
SQLStateAVG()
SQLCountryAVG()
Diff()

wb.save(filename = dest_filename)
conn.close()
