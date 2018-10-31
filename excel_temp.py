import openpyxl as pyxl

import sqlite3 as sql3

from openpyxl.chart import LineChart, Reference, Series, AreaChart

conn = sql3.connect("data/collectedData.db")#open database
cur = conn.cursor()

wb = pyxl.Workbook()#open workbook

dest_filename = 'WorldTemperature.xlsx'
#create sheet for data
ws = wb.active
ws.title = "TempByCity"
sheet = wb.get_sheet_by_name('TempByCity')


#show all cities in China
taks1 = cur.execute("Select distinct City from City where Country == 'China';")# select all cities in china
City = []
for row in taks1:
   print("City = ", row[0],"\n")
   City.append(row[0])#save cities
   #print("City = ", row[1],)
   #print("AVG = ", row[2],"\n")

sheet['B1'] = 'AVG Yearly Temp'
sheet['A1'] = 'City'
count = 2
for i in City:#find all cities that match criteria
    taks1 = cur.execute("Select avg(AverageTemp) as Temp, City from City where Date Like '_____01_01' and Country == 'China' and City == '"+i+"';")
    for row in taks1:
       #print("Date = ", row[0],)
       print("Avg Yearly Temp= ", row[0],)
       print("City = ", row[1],"\n")
       sheet.cell(row = count, column = 2, value = row[0]) 
       sheet.cell(row = count, column = 1, value = row[1])
       count += 1


#Line Chart
#generate chart
data = Reference(ws, min_col=2, min_row=1, max_col=2, max_row=count)
titles = Reference(ws, min_col=1, min_row=2, max_row=count)
chart = LineChart()
chart.title = "Average Yearly Temp Per city in China"
chart.add_data(data=data, titles_from_data=True)
chart.set_categories(titles)
chart.y_axis.title = 'Temp C'
chart.x_axis.title = 'City'

ws.add_chart(chart, "E15")


wb.save(filename = dest_filename)

conn.commit()#save database
conn.close()#close database
