import openpyxl as pyxl

from threading import Thread
 
import sqlite3 as sql3

conn = sql3.connect("data/collectedData.db")#load database
cur = conn.cursor()

#state global var
wbCity = None 
wbState = None
wbCountry = None

def loadWS():
    global wbCity,wbState,wbCountry
    
    #create exel reader
    wbCity = pyxl.load_workbook("data/GlobalLandTemperaturesByMajorCity.xlsx")
    cityName = wbCity.get_sheet_names()
    print("sheet 1 got")
    
    wbState = pyxl.load_workbook("data/GlobalLandTemperaturesByState.xlsx")
    stateName = wbState.get_sheet_names()
    print("sheet 2 got")
    
    wbCountry = pyxl.load_workbook("data/GlobalLandTemperaturesByCountry.xlsx")
    countryName = wbCountry.get_sheet_names()
    print("sheet 3 got")

    name = [cityName, stateName, countryName]
    print(name)



def CreateDatabase():
    #create database with 3 tables
    '''
    Create Table "Southern cities"
    '''
    try:
        sql_dropcommand = "drop table State"
        sql_dropcommand1 = "drop table Country"
        sql_dropcommand2 = "drop table City"
        sql_dropcommand3 = "drop table SouthernCities"
        cur.execute(sql_dropcommand)
        cur.execute(sql_dropcommand1)
        cur.execute(sql_dropcommand2)
        cur.execute(sql_dropcommand3)
    except:
        print()
    #ID INTEGER PRIMARY KEY, 
    sql_command = """
    CREATE TABLE State (
    Date CHAR(30),
    AverageTemp int(30), 
    AvgTempUncertain int(20),
    State CHAR(20),
    Country CHAR(20));
    """


    sql_command1 = """
    CREATE TABLE Country ( 
    Date CHAR(30), 
    AverageTemp int(30), 
    AvgTempUncertain int(20), 
    Country CHAR(20));
    """


    sql_command2 = """
    CREATE TABLE City (
    Date CHAR(30),
    AverageTemp VARCHAR(20), 
    AvgTempUncertain VARCHAR(30), 
    City CHAR(20), 
    Country CHAR(20),
    Lat CHAR(20),
    Long CHAR(20));
    """

    sql_command3 = """
    CREATE TABLE SouthernCities (
    City CHAR(20), 
    Country CHAR(20),
    Lat CHAR(20),
    Long CHAR(20));
    """

    cur.execute(sql_command)
    cur.execute(sql_command1)
    cur.execute(sql_command2)
    cur.execute(sql_command3)



    print("Added Table")
    conn.commit()

class imports:
    global wbCity,wbState,wbCountry
    def City():
        #add data from exel
        #CITY
        print("CITY")
        ws=wbCity.active

        count = 2

        while True: #cell_1 != None:
            Value = []
            first = None
            for i in [1,2,3,4,5,6,7]: #collect row data
                #cell_1 = ws[i+count]
                Score = ws.cell(row=count, column=i).value
                if count == 2 and i == 1:#loop fail safe
                    first == Score
                elif i == 1 and Score == first:
                    False
                    break
                Value.append(Score) 
            try:
                cur.execute("INSERT INTO City VALUES ( ?, ?, ?, ?, ?, ?, ?);", (Value[0], Value[1], Value[2], Value[3],Value[4],Value[5],Value[6]))
                #conn.commit()
            except:
                print("CITY FIN")
                break
                
                #print(cell_1.value)
            count += 1

    def Country():
        #COUNTRY
        print("COUNTRY")
        ws=wbCountry.active

        count = 2

        while True:
            Value = []
            first = None
            for i in [1,2,3,4]: #collect row data
                Score = ws.cell(row=count, column=i).value
                if count == 2 and i == 1:#loop fail safe
                    first == Score
                elif i == 1 and Score == first:
                    False
                    break
                Value.append(Score) 
            try:
                cur.execute("INSERT INTO Country VALUES ( ?, ?, ?, ?);", (Value[0], Value[1], Value[2], Value[3]))
                #conn.commit()
            except:
                print("COUNTRY FIN")
                break
            #if Value == [None,None,None,None]:
                #break
            count += 1

    def State():
        #STATE
        print("STATE")
        ws=wbState.active

        count = 2

        while True:
            Value = []
            first = None
            for i in [1,2,3,4,5]: #collect row data
                Score = ws.cell(row=count, column=i).value
                if count == 2 and i == 1:#loop fail safe
                    first == Score
                elif i == 1 and Score == first:
                    False
                    break
                Value.append(Score) 
            try:
                cur.execute("INSERT INTO State VALUES ( ?, ?, ?, ?, ?);", (Value[0], Value[1], Value[2], Value[3],Value[4]))
                #conn.commit()
            except:
                print("STATE FIN")
                break
            count += 1


#end

if __name__ == "__main__":
    loadWS()
    CreateDatabase()

    #load data via miltithread
    P1 = Thread(target=imports.City())
    P2 = Thread(target=imports.Country())
    P3 = Thread(target=imports.State())
    
    P1.start()
    P2.start()
    P3.start()
    
    P1.join()
    P2.join()
    P3.join()

    print("done")
    
    conn.commit()#save database

    conn.close()#close database
